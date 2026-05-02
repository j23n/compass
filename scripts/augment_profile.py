# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl"]
# ///
"""
augment_profile.py — Merge official FIT SDK profile with Gadgetbridge additions.

Priority: Official SDK (never overwritten) > Gadgetbridge

Prerequisites:
  1. Copy Profile.xlsx from the roznet SPM checkout:
       swift build in Packages/CompassFIT to resolve the dependency, then:
       cp .build/checkouts/FitFileParser-*/python/Profile.xlsx data/fit-sdk/Profile.xlsx
       (Xcode: find in ~/Library/Developer/Xcode/DerivedData/*/SourcePackages/checkouts/)
  2. Ensure data/gadgetbridge/NativeFITMessage.java is present (already fetched).

Usage:
    uv run python scripts/augment_profile.py \\
      --base    data/fit-sdk/Profile.xlsx \\
      --gadget  data/gadgetbridge/NativeFITMessage.java \\
      --out     data/fit-sdk/Profile.aug.xlsx \\
      --conflicts data/conflicts.md

After reviewing data/conflicts.md, run fitsdkparser.py to regenerate Swift:
    cd Packages/FitFileParser/python
    uv run python fitsdkparser.py generate ../../../data/fit-sdk/Profile.aug.xlsx \\
      -o ../Sources/FitFileParserObjc \\
      -s ../Sources/FitFileParser
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Column indices in the Messages sheet (0-based), matching fitsdkparser.py
# ---------------------------------------------------------------------------
MSG_COL_MESSAGE_NAME    = 0
MSG_COL_FIELD_DEF_NUM   = 1
MSG_COL_FIELD_NAME      = 2
MSG_COL_FIELD_TYPE      = 3
MSG_COL_ARRAY           = 4
MSG_COL_COMPONENTS      = 5
MSG_COL_SCALE           = 6
MSG_COL_OFFSET          = 7
MSG_COL_UNITS           = 8
MSG_COL_BITS            = 9
MSG_COL_ACCUMULATE      = 10
MSG_COL_REF_FIELD_NAME  = 11
MSG_COL_REF_FIELD_VALUE = 12
MSG_COL_COMMENT         = 13
MSG_COL_PRODUCTS        = 14
MSG_COL_EXAMPLE         = 15
MSG_TOTAL_COLS          = 16


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class FieldDef:
    field_def_num: int
    field_name: str
    field_type: str = ""
    array: str = ""
    components: str = ""
    scale: str = ""
    offset: str = ""
    units: str = ""
    bits: str = ""
    accumulate: str = ""
    ref_field_name: str = ""
    ref_field_value: str = ""
    comment: str = ""
    products: str = ""
    example: str = ""
    source: str = ""  # "official" | "gadgetbridge"


@dataclass
class MessageDef:
    message_name: str
    global_message_number: Optional[int]
    fields: dict[int, FieldDef] = field(default_factory=dict)
    source: str = ""  # "official" | "gadgetbridge"


# ---------------------------------------------------------------------------
# BaseType → FIT type string
# ---------------------------------------------------------------------------

_BASETYPE_MAP = {
    "ENUM":           "enum",
    "SINT8":          "sint8",
    "UINT8":          "uint8",
    "SINT16":         "sint16",
    "UINT16":         "uint16",
    "SINT32":         "sint32",
    "UINT32":         "uint32",
    "FLOAT32":        "float32",
    "FLOAT64":        "float64",
    "UINT8Z":         "uint8z",
    "UINT16Z":        "uint16z",
    "UINT32Z":        "uint32z",
    "BYTE":           "byte",
    "BASE_TYPE_BYTE": "byte",
    "SINT64":         "sint64",
    "UINT64":         "uint64",
    "UINT64Z":        "uint64z",
    "STRING":         "string",
}


def _basetype_to_fit(java_type: str) -> str:
    return _BASETYPE_MAP.get(java_type.upper(), java_type.lower())


# ---------------------------------------------------------------------------
# Gadgetbridge NativeFITMessage.java parser
# ---------------------------------------------------------------------------

# Match: new NativeFITMessage(273, "SLEEP_DATA_INFO", Arrays.asList(
_MSG_CTOR = re.compile(
    r'new\s+NativeFITMessage\s*\(\s*(\d+)\s*,\s*"([^"]+)"\s*,\s*Arrays\.asList\s*\(',
)

# Match: new FieldDefinitionPrimitive(fieldNum, BaseType.TYPE, "fieldName", ...)
_FIELD_PRIM = re.compile(
    r'new\s+FieldDefinitionPrimitive\s*\(\s*'
    r'(\d+)'                          # group 1: field number
    r'\s*,\s*BaseType\.(\w+)'         # group 2: base type
    r'\s*,\s*"([^"]*)"'               # group 3: field name
    r'(?:\s*,\s*([^)]+))?'            # group 4: optional extra args (scale, offset, factory)
    r'\s*\)',
)


def _parse_extra_args(extra: str | None) -> tuple[str, str]:
    """Extract (scale, offset) from the optional extra args string, if numeric."""
    if not extra:
        return "", ""
    parts = [p.strip() for p in extra.split(",")]
    scale, offset = "", ""
    for i, p in enumerate(parts):
        try:
            float(p)
            if i == 0:
                scale = p.strip()
            elif i == 1:
                offset = p.strip()
        except ValueError:
            pass
    return scale, offset


def parse_gadgetbridge(java_path: Path) -> dict[int, MessageDef]:
    """Parse NativeFITMessage.java → dict of msg_num → MessageDef."""
    text = java_path.read_text(encoding="utf-8")
    messages: dict[int, MessageDef] = {}

    msg_starts = list(_MSG_CTOR.finditer(text))
    for i, m in enumerate(msg_starts):
        msg_num  = int(m.group(1))
        msg_name = m.group(2).lower()

        block_start = m.end()
        block_end = len(text) if i + 1 >= len(msg_starts) else msg_starts[i + 1].start()
        block = text[block_start:block_end]

        msg_def = MessageDef(
            message_name=msg_name,
            global_message_number=msg_num,
            source="gadgetbridge",
        )
        for f in _FIELD_PRIM.finditer(block):
            fnum  = int(f.group(1))
            ftype = _basetype_to_fit(f.group(2))
            fname = f.group(3)
            scale, offset = _parse_extra_args(f.group(4))
            msg_def.fields[fnum] = FieldDef(
                field_def_num=fnum,
                field_name=fname,
                field_type=ftype,
                scale=scale,
                offset=offset,
                source="gadgetbridge",
            )

        messages[msg_num] = msg_def

    return messages


# ---------------------------------------------------------------------------
# Official Profile.xlsx loader
# ---------------------------------------------------------------------------

def load_official_profile(xlsx_path: Path) -> dict[int, MessageDef]:
    """Load the official FIT SDK Profile.xlsx Messages sheet.

    Returns dict of msg_num → MessageDef (negative sentinels for unnumbered messages).
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb["Messages"]
    except KeyError:
        raise ValueError(f"No 'Messages' sheet in {xlsx_path}; found: {wb.sheetnames}")

    messages: dict[int, MessageDef] = {}
    current_msg: Optional[MessageDef] = None
    _sentinel = -1

    for row in list(ws.iter_rows(values_only=True))[1:]:
        cells = [str(c).strip() if c is not None else "" for c in row]
        if len(cells) < MSG_TOTAL_COLS:
            cells += [""] * (MSG_TOTAL_COLS - len(cells))

        msg_name  = cells[MSG_COL_MESSAGE_NAME]
        field_num = cells[MSG_COL_FIELD_DEF_NUM]

        if msg_name and not field_num:
            _sentinel -= 1
            current_msg = MessageDef(
                message_name=msg_name.lower(),
                global_message_number=None,
                source="official",
            )
            messages[_sentinel] = current_msg
            continue

        if current_msg is None:
            continue

        try:
            fnum = int(field_num)
        except (ValueError, TypeError):
            continue

        current_msg.fields[fnum] = FieldDef(
            field_def_num=fnum,
            field_name=cells[MSG_COL_FIELD_NAME],
            field_type=cells[MSG_COL_FIELD_TYPE],
            array=cells[MSG_COL_ARRAY],
            components=cells[MSG_COL_COMPONENTS],
            scale=cells[MSG_COL_SCALE],
            offset=cells[MSG_COL_OFFSET],
            units=cells[MSG_COL_UNITS],
            bits=cells[MSG_COL_BITS],
            accumulate=cells[MSG_COL_ACCUMULATE],
            ref_field_name=cells[MSG_COL_REF_FIELD_NAME],
            ref_field_value=cells[MSG_COL_REF_FIELD_VALUE],
            comment=cells[MSG_COL_COMMENT],
            products=cells[MSG_COL_PRODUCTS],
            example=cells[MSG_COL_EXAMPLE],
            source="official",
        )

    # Resolve global message numbers from the Types / mesg_num block
    name_to_def: dict[str, MessageDef] = {d.message_name.lower(): d for d in messages.values()}
    try:
        nums_ws = wb["Types"]
        current_type = ""
        for row in list(nums_ws.iter_rows(values_only=True))[1:]:
            cells2 = [str(c).strip() if c is not None else "" for c in row]
            if len(cells2) < 4:
                continue
            if cells2[0]:
                current_type = cells2[0].lower()
            if current_type == "mesg_num" and cells2[2]:
                val_name = cells2[2].lower()
                try:
                    num = int(cells2[3])
                    if val_name in name_to_def:
                        name_to_def[val_name].global_message_number = num
                except (ValueError, TypeError):
                    pass
    except KeyError:
        pass

    wb.close()

    result: dict[int, MessageDef] = {}
    unnamed_counter = -1
    for msg in messages.values():
        if msg.global_message_number is not None and msg.global_message_number >= 0:
            result[msg.global_message_number] = msg
        else:
            result[unnamed_counter] = msg
            unnamed_counter -= 1
    return result


# ---------------------------------------------------------------------------
# Merge logic
# ---------------------------------------------------------------------------

@dataclass
class ConflictEntry:
    msg_num: int
    msg_name: str
    field_num: Optional[int]   # None = message-level conflict
    description: str
    official_val: str          # "" if official has no entry
    gadgetbridge_val: str      # "" if gadgetbridge has no entry
    resolution: str = ""       # auto-applied resolution; empty = needs review


def merge(
    official: dict[int, MessageDef],
    gadgetbridge: dict[int, MessageDef],
) -> tuple[dict[int, MessageDef], list[ConflictEntry]]:
    """Merge official SDK and Gadgetbridge with priority official > gadgetbridge."""
    merged: dict[int, MessageDef] = {}
    conflicts: list[ConflictEntry] = []

    official_nums = set(n for n in official if n >= 0)
    official_names_lower = {d.message_name.lower() for d in official.values()}

    # Seed with official (never overwritten)
    for num, msg in official.items():
        merged[num] = msg

    # Merge Gadgetbridge
    for msg_num, gbd_msg in gadgetbridge.items():
        if msg_num in official_nums:
            # Message exists in official — merge fields only; official fields are immutable
            existing = merged[msg_num]
            for fnum, fdef in gbd_msg.fields.items():
                if fnum not in existing.fields:
                    existing.fields[fnum] = fdef
                # official fields always win; no conflict needed for field-level
            # Flag name mismatch at message level
            if existing.message_name.lower() != gbd_msg.message_name.lower():
                conflicts.append(ConflictEntry(
                    msg_num=msg_num,
                    msg_name=existing.message_name,
                    field_num=None,
                    description="Sources use different names for this message number",
                    official_val=existing.message_name,
                    gadgetbridge_val=gbd_msg.message_name,
                    resolution="Official SDK (auto-applied)",
                ))
        else:
            # New Gadgetbridge-only message — check for name collision
            if gbd_msg.message_name in official_names_lower:
                off_num = next(
                    (n for n, m in official.items() if m.message_name.lower() == gbd_msg.message_name),
                    "?",
                )
                conflicts.append(ConflictEntry(
                    msg_num=msg_num,
                    msg_name=gbd_msg.message_name,
                    field_num=None,
                    description="Gadgetbridge assigns a different message number to a name already in Official SDK",
                    official_val=f"msg_num={off_num}",
                    gadgetbridge_val=f"msg_num={msg_num}",
                    resolution="Official SDK (auto-applied)",
                ))
            merged[msg_num] = gbd_msg

    return merged, conflicts


# ---------------------------------------------------------------------------
# xlsx writer
# ---------------------------------------------------------------------------

def _add_mesg_num_entries(ws_types, merged: dict[int, MessageDef]) -> None:
    """Insert mesg_num type entries for every non-official message into the mesg_num block.

    openpyxl can't insert mid-sheet, so we read all rows, inject the new entries into the
    mesg_num block, clear the sheet, and rewrite.
    """
    all_rows = list(ws_types.iter_rows(values_only=True))
    header_row = all_rows[0]

    # Parse existing mesg_num values and locate the block boundaries
    mesg_num_existing: set[str] = set()
    mesg_num_end_idx: int | None = None   # index (in all_rows) of first row AFTER the mesg_num block

    in_mesg_num = False
    for i, row in enumerate(all_rows[1:], start=1):
        cells = [str(c).strip() if c is not None else "" for c in row]
        is_type_header = bool(cells[0] and cells[1])
        if is_type_header and cells[0] == "mesg_num":
            in_mesg_num = True
        elif is_type_header and in_mesg_num:
            mesg_num_end_idx = i
            in_mesg_num = False
        if in_mesg_num and cells[2]:
            mesg_num_existing.add(cells[2])

    # Build new entries
    new_entries: list[tuple] = []
    for msg_num in sorted(k for k in merged if k >= 0):
        msg = merged[msg_num]
        if msg.source == "official":
            continue
        if msg.message_name not in mesg_num_existing:
            new_entries.append((None, None, msg.message_name, msg_num, msg.source))
            mesg_num_existing.add(msg.message_name)

    if not new_entries:
        return

    # Inject new entries into the mesg_num block, then rewrite the sheet
    insert_at = mesg_num_end_idx if mesg_num_end_idx is not None else len(all_rows)
    rebuilt = all_rows[:insert_at] + new_entries + all_rows[insert_at:]

    ws_types.delete_rows(1, ws_types.max_row + 1)
    ws_types.append(list(header_row))
    for row in rebuilt[1:]:
        ws_types.append(list(row))


def write_augmented_xlsx(
    merged: dict[int, MessageDef],
    official_wb_path: Path,
    out_path: Path,
) -> None:
    """Copy the official xlsx, append Gadgetbridge messages, extend mesg_num type."""
    import shutil
    shutil.copy2(official_wb_path, out_path)

    wb = load_workbook(out_path)
    ws_msg   = wb["Messages"]
    ws_types = wb["Types"] if "Types" in wb.sheetnames else wb.create_sheet("Types")

    _add_mesg_num_entries(ws_types, merged)

    # Collect message names already written in the official base
    official_names_in_xlsx: set[str] = set()
    for row in ws_msg.iter_rows(min_row=2, values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if cells[0] and not cells[1]:
            official_names_in_xlsx.add(cells[0].lower())

    # Append non-official messages
    for msg_num in sorted(k for k in merged if k >= 0):
        msg = merged[msg_num]
        if msg.source == "official" or msg.message_name.lower() in official_names_in_xlsx:
            continue

        header = [""] * MSG_TOTAL_COLS
        header[MSG_COL_MESSAGE_NAME] = msg.message_name
        ws_msg.append(header)

        for fnum in sorted(msg.fields):
            f = msg.fields[fnum]
            field_row = [""] * MSG_TOTAL_COLS
            field_row[MSG_COL_FIELD_DEF_NUM] = fnum
            field_row[MSG_COL_FIELD_NAME]    = f.field_name
            field_row[MSG_COL_FIELD_TYPE]    = f.field_type
            field_row[MSG_COL_ARRAY]         = f.array
            field_row[MSG_COL_SCALE]         = f.scale
            field_row[MSG_COL_OFFSET]        = f.offset
            field_row[MSG_COL_UNITS]         = f.units
            field_row[MSG_COL_COMMENT]       = f.comment
            ws_msg.append(field_row)

    wb.save(out_path)
    wb.close()
    print(f"Wrote {out_path}")


# ---------------------------------------------------------------------------
# conflicts.md writer
# ---------------------------------------------------------------------------

_KNOWN_PARSER_NOTES: dict[tuple[int, Optional[int]], str] = {
    (140, None): "Our `MonitoringFITParser` dispatched on msg 140 as `monitoring_hr` — field 1 read as per-sample HR. Actually `PHYSIOLOGICAL_METRICS` (post-activity summary); field 1 = `new_hr_max`. **Fix: remove msg 140 dispatch from MonitoringFITParser.**",
    (346, None): "Our `MonitoringFITParser` dispatched on msg 346 as `body_battery`. Actually `SLEEP_STATS` (sleep quality sub-scores). Body battery comes from `hsa_body_battery_data` (314), already correct. **Fix: remove msg 346 dispatch from MonitoringFITParser.**",
    (273, 1):    "Our `SleepFITParser` reads field 1 as `sleep_score` (uint8, 0–100). Gadgetbridge names it `sample_length` (uint16, typically 60 = 60 s). We may be reporting the sample interval as the sleep score for every night.",
    (274, 0):    "Our `SleepFITParser` reads field 0 as uint8 sleep level. Gadgetbridge says 20 raw bytes per sample. First byte may coincidentally match level encoding on Instinct Solar 1G.",
    (275, 1):    "Our `SleepFITParser` reads field 1 as `duration`. Gadgetbridge has no field 1 in SLEEP_STAGE.",
    (276, None): "Our `SleepFITParser` field-dumps msg 276. No Gadgetbridge entry. Should redirect to msg 346 (`SLEEP_STATS`) for quality sub-scores.",
    (382, 0):    "Our `SleepFITParser` logs msg 382 existence only. Gadgetbridge has field 0 = `unknown_0`, field 1 = `restless_moments_count`, field 2 = `durations`.",
}


def write_conflicts_md(conflicts: list[ConflictEntry], out_path: Path) -> None:
    lines = [
        "# FIT Profile Augmentation — Conflict Report\n",
        "Generated by `scripts/augment_profile.py` (sources: Official SDK + Gadgetbridge).\n",
        "All conflicts listed here are auto-resolved by the priority rule: Official SDK > Gadgetbridge.\n",
        "",
        "## Confirmed Parser Errors\n",
        "",
    ]
    for (msg_num, field_num), note in _KNOWN_PARSER_NOTES.items():
        key = f"Msg {msg_num}" + (f" field {field_num}" if field_num is not None else "")
        lines.append(f"### {key}\n")
        lines.append(note + "\n")
        lines.append("")

    lines += ["---\n", "## Source-Comparison Conflicts\n", ""]

    if not conflicts:
        lines.append("_No source-comparison conflicts detected._\n")
    else:
        for c in conflicts:
            field_label = f"field {c.field_num}" if c.field_num is not None else "message level"
            lines.append(f"### Message {c.msg_num} `{c.msg_name}` — {field_label}\n")
            lines.append(f"**Issue:** {c.description}\n")
            lines.append("")
            lines.append("| Source | Value |")
            lines.append("|---|---|")
            lines.append(f"| Official SDK | `{c.official_val}` |" if c.official_val else "| Official SDK | _(not present)_ |")
            lines.append(f"| Gadgetbridge | `{c.gadgetbridge_val}` |" if c.gadgetbridge_val else "| Gadgetbridge | _(not present)_ |")
            lines.append("")
            if c.resolution:
                lines.append(f"**Resolution:** {c.resolution}\n")
            else:
                lines.append("**Decision needed:** (fill in after reviewing live FIT data)\n")
            lines.append("")

    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {out_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Augment FIT SDK Profile.xlsx with Gadgetbridge additions"
    )
    parser.add_argument("--base",      required=True, type=Path, help="Official Profile.xlsx from roznet")
    parser.add_argument("--gadget",    required=True, type=Path, help="NativeFITMessage.java from Gadgetbridge")
    parser.add_argument("--out",       required=True, type=Path, help="Output Profile.aug.xlsx")
    parser.add_argument("--conflicts", required=True, type=Path, help="Output conflicts.md")
    args = parser.parse_args()

    for p, label in [(args.base, "--base"), (args.gadget, "--gadget")]:
        if not p.exists():
            print(f"ERROR: {label} file not found: {p}", file=sys.stderr)
            if label == "--base":
                print("  Hint: run `swift build` in Packages/CompassFIT, then:", file=sys.stderr)
                print("    cp .build/checkouts/FitFileParser-*/python/Profile.xlsx data/fit-sdk/Profile.xlsx", file=sys.stderr)
            sys.exit(1)

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.conflicts.parent.mkdir(parents=True, exist_ok=True)

    print("Loading official Profile.xlsx …")
    official = load_official_profile(args.base)
    print(f"  {len(official)} messages (including sentinel entries)")

    print("Parsing Gadgetbridge NativeFITMessage.java …")
    gadgetbridge = parse_gadgetbridge(args.gadget)
    print(f"  {len(gadgetbridge)} messages")

    print("Merging …")
    merged, conflicts = merge(official, gadgetbridge)
    print(f"  {len(merged)} total messages, {len(conflicts)} conflicts")

    write_augmented_xlsx(merged, args.base, args.out)
    write_conflicts_md(conflicts, args.conflicts)

    print("\nNext steps:")
    print("  Run fitsdkparser.py to regenerate Swift:")
    print("    cd Packages/FitFileParser/python")
    print("    uv run python fitsdkparser.py generate ../../../data/fit-sdk/Profile.aug.xlsx \\")
    print("      -o ../Sources/FitFileParserObjc -s ../Sources/FitFileParser")


if __name__ == "__main__":
    main()
